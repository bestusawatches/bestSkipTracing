from fastapi import APIRouter, Header, File, Request, UploadFile, Form, Body
from typing import Any
import time
import re
import json
import openpyxl
from zenrows import ZenRowsClient
from bs4 import BeautifulSoup
import asyncio

router = APIRouter()
# zenrows_key = "e1871ef65e65d4d4cb01549a7d53e616adc227cb"
zenrows_key = "28bbe7ae1b3c6541a1b32891685ffb17f8a43df7"

def read_xlsx(file):
    wb = openpyxl.load_workbook(file)
    worksheet = wb.active
    keys = []
    data = []
    for row in worksheet.iter_rows():
        if len(keys):
            line = {}
            i = 0
            for cell in row:
                line[keys[i]] = cell.value
                i += 1
            data.append(line)
        else:
            for cell in row:
                keys.append(cell.value)
    return data

def convert_data(data1):
    # data2 = json.dumps([dict(zip(map(lambda x:x['title'], data1["header"]), row)) for row in data1["data"]], separators=(',', ':'))
    # data2_dict = json.loads(data2)
    # data2_dict = [{'File #': 1, 'Address': '5441 SKYCREST DRIVE.', 'City': 'Ames, IA 50010', 'Age1': None, 'Phone1-1': None, 'Phone1-2': None, 'Phone1-3': None, 'Phone1-4': None, 'Phone1-5': None, 'Phone1-6': None, 'Email1-1': None, 'Email1-2': None}, {'File #': 2, 'Address': '32825 710TH AVE.', 'City': 'Collins, IA 50055', 'Age1': None, 'Phone1-1': None, 'Phone1-2': None, 'Phone1-3': None, 'Phone1-4': None, 'Phone1-5': None, 'Phone1-6': None, 'Email1-1': None, 'Email1-2': None}]
    result = []
    for row in data1['data']:
        obj = {}
        for i, header in enumerate(data1['header']):
            obj[header['title']] = row[i] or None
        result.append(obj)
    
    print(data1)
    print(result)
    print(type(result))
    return result

def write_xlsx(file, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, key in enumerate(data[0].keys(), start=1):
        ws.cell(row=1, column=col, value=key)

    for row, item in enumerate(data, start=2):
        for col, value in enumerate(item.values(), start=1):
            ws.cell(row=row, column=col, value=value)
    
    wb.save(file)


def build_url(data, owner_num):
    Address = data["Address"]
    City = data["City"]
    State = data["State"]
    Zipcode = data["Zipcode"]
    search_url = "https://www.truepeoplesearch.com/resultaddress?streetaddress=" + Address.replace(" ", "%20").replace(",", "") + "&citystatezip=" + City.replace(" ", "%20").replace(",", "") + "%20" + State.replace(",", "") + "%20" + str(Zipcode).replace(",", "")

    return search_url


async def scrap(input_data, client, params):

    data = {
    **input_data,
    'Age1': '',
    'Phone-1-Number': '',
    'Phone-1-Type': '',
    'Phone-2-Number': '',
    'Phone-2-Type': '',
    'Phone-3-Number': '',
    'Phone-3-Type': '',
    'Phone-4-Number': '',
    'Phone-4-Type': '',
    'Phone-5-Number': '',
    'Phone-5-Type': '',
    'Phone-6-Number': '',
    'Phone-6-Type': '',
    'Email1-1': '',
    'Email1-2': '',
    'First Name': '',
    'Last Name': ''
    }

    Address = data["Address"]
    City = data["City"]
    State = data["State"]
    Zipcode = data["Zipcode"]
    
    if(Address != None and City != None and State != None and Zipcode != None):
    
        result = []
        age_regex = re.compile(r'Age \d+ \([A-Za-z]{3} \d+\)')

        search_url = build_url(data, 1)

        # print("\n")
        # print(search_url)

        response = await client.get_async(search_url, params=params)

        with open("home_1.txt", "w", encoding="UTF-8") as output:
            output.write(response.text)

        soup = BeautifulSoup(response.text, 'html.parser')

        # state = data["City"].split(",")[0].lower()
        state = data["City"] + " " + data["State"] + " " + str(data["Zipcode"])
        # print(state)

        link = ""

        for div in soup.find_all('div', {'class': 'card-summary'}):
            texts = div.text.split("\n")
            new_texts = [item.strip() for item in texts if item != '']
            
            ind = new_texts.index('Lives in')
            if state in new_texts[ind+1].lower():
                link = div.find('a', {'class': 'detail-link'})['href']
                break

        try:
            if link == "":
                link = soup.find_all('a', {'class': 'btn btn-success btn-lg detail-link shadow-form'})
            # print(len(link))
            for ind in range(0, len(link), 2):
                data = {
                    **input_data,
                    'Age1': '',
                    'Phone-1-Number': '',
                    'Phone-1-Type': '',
                    'Phone-2-Number': '',
                    'Phone-2-Type': '',
                    'Phone-3-Number': '',
                    'Phone-3-Type': '',
                    'Phone-4-Number': '',
                    'Phone-4-Type': '',
                    'Phone-5-Number': '',
                    'Phone-5-Type': '',
                    'Phone-6-Number': '',
                    'Phone-6-Type': '',
                    'Email1-1': '',
                    'Email1-2': '',
                    'First Name': '',
                    'Last Name': ''
                }
                # print(link[ind]['href'])
                detail_url = "https://www.truepeoplesearch.com" + link[ind]['href']

                # print(detail_url)

                detail_response = await client.get_async(detail_url, params=params)

                with open(f"detail_{ind}.txt", "w", encoding="UTF-8") as output:
                    output.write(detail_response.text)

                detail_soup = BeautifulSoup(detail_response.text, 'html.parser')
                
                owner = detail_soup.select_one('h1.oh1').text
                name = owner.split(" ")
                data["First Name"] = name[0]
                data["Last Name"] = name[len(name) -1]
                
                try:
                    age_span_1 = detail_soup.find('span', string=age_regex)    
                    data["Age1"] = age_span_1.text.strip()
                except:
                    data["Age1"] = "Age Unknown"
                
                phone_numbers_1 = []
                phone_divs = detail_soup.select('div.row div.col-12')
                for phone_div in phone_divs:
                    try:
                        phone_span = phone_div.select_one('a[data-link-to-more="phone"] span[itemprop="telephone"]')
                        phone = phone_span.text.strip()

                        phone_type_span = phone_div.select_one('span.smaller')
                        phone_type = phone_type_span.text.strip()

                        phone_info = f"{phone} - {phone_type}"
                        if phone_info not in phone_numbers_1:
                            phone_numbers_1.append(phone_info)
                    except:
                        pass
                i = 1
                
                # print(len(phone_numbers_1))
                row_enable = len(phone_numbers_1)
                while len(phone_numbers_1):
                    phone = phone_numbers_1.pop(0)
                    phone_number = re.sub('[^0-9]', '', phone)
                    phone_type = phone.split(' - ')[1]
                    
                    if i < 7:
                        data["Phone-"+str(i)+"-Number"] = int(phone_number)
                        data["Phone-"+str(i)+"-Type"] = phone_type
                        i += 1
                # print(data)

                emails_1 = []
                email_divs = detail_soup.select('.col > div:last-child')
                for email_div in email_divs:
                    text = email_div.text.strip()
                    if re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text):
                        if text != "support@truepeoplesearch.com" and text not in emails_1:
                            emails_1.append(text)

                j = 1
                while len(emails_1):
                    email = emails_1.pop(0)
                    if j < 3:
                        data["Email1-"+str(j)] = email
                        j += 1
                if(row_enable > 0):
                    print('ok')
                    result.append(data.copy())
        except:
            print("Cannot Find Detail Link")
        return result


async def main(data):
    client = ZenRowsClient(zenrows_key, concurrency=25, retries=1)
    params = {"js_render":"true","antibot":"false","premium_proxy":"true"}

    # input_file = 'Test2.xlsx'
    # input_data = read_xlsx(input_file)

    input_data = convert_data(data)

    # Prompt for the range of rows to scrape
    # start_row = int(input("Enter the starting row (e.g., 1): "))
    # end_row = int(input("Enter the ending row (e.g., 20): "))

    # Slice the input_data based on the specified range
    # input_temp = input_data[start_row - 1:end_row]

    input_temp = input_data

    tasks = [scrap(item, client, params) for item in input_temp]
    output_1 = await asyncio.gather(*tasks)
    output_data = []

    for sublist  in output_1:
        # print(sublist)
        if(sublist != None):
            for item in sublist:
                output_data.append(item)

    # output_file = f"result.xlsx"
    
    # write_xlsx(output_file, output_data)

    return output_data

@router.post('/')
async def tracing_demo(req: Request):
    body = await req.body()
    data = json.loads(body)
    try:
        # result = asyncio.create_task(main(data))
        result = await asyncio.create_task(main(data))
        # input_data = convert_data(data)
        return{
        "Success": True,
        "data": result
        }
    except Exception as e:
       print("Error occurred:", str(e))
    # return{
    #     "Success": True,
    #     "data": [{'File #': 1, 'Mailing Address': '139 BAY STREET', 'Mailing City St  Zip': 'STATEN ISLAND NY 10301', 'Owner name 01': '139 BAY POINTE PROPERTIES LLC', 'Owner name 02': '', "Entity Owner's Name (Only if ENTITY)": 'Jeanette Baldassano', 'Age1': 'Age 72 (Aug 1951)', 'Phone1-1': '(917) 474-6724 - Wireless', 'Phone1-2': '(718) 783-8789 - Landline', 'Phone1-3': '(718) 832-8057 - Landline', 'Phone1-4': '(607) 498-5311 - Landline', 'Phone1-5': '(718) 981-5817 - Landline', 'Phone1-6': '(718) 981-5316 - Landline', 'Email1-1': 'blennyb@aol.com'}, {'File #': '2', 'Mailing Address': '15 PAGE AVE', 'Mailing City St  Zip': 'STATEN ISLAND NY 10309-2611', 'Owner name 01': '139 BAY STREET REALTY LLC', 'Owner name 02': '', "Entity Owner's Name (Only if ENTITY)": 'Frank G Damato', 'Age1': 'Age 62 (Aug 1961)', 'Phone1-1': '(646) 808-6460 - Wireless', 'Phone1-2': '(718) 967-8393 - Landline', 'Phone1-3': '(276) 546-4274 - Landline', 'Phone1-4': '(732) 202-7154 - VOIP', 'Phone1-5': '(718) 763-9103 - Landline', 'Phone1-6': '(718) 967-0000 - Landline', 'Email1-1': 'fgd4712@aol.com'}]
    # }