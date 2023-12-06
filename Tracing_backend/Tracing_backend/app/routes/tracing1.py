from fastapi import APIRouter, Header, File, Request, UploadFile, Form, Body
from typing import Any
import time
import re
import json
import openpyxl
from zenrows import ZenRowsClient
from bs4 import BeautifulSoup
import asyncio

# zenrows_key = "96e9e9b7c8e73a277886ea5bd93e2cb07ee069d7"
zenrows_key = "c8d5abff98b190cdbb19421eafee73be01eb3bd3"

router = APIRouter()

def read_xlsx(file):
    wb = openpyxl.load_workbook(file)
    worksheet = wb.active
    keys = []
    data = []
    for row in worksheet.iter_rows():
        if len(keys):
            line = {}
            i = 0
            for cell in row:
                line[keys[i]] = cell.value
                i += 1
            data.append(line)
        else:
            for cell in row:
                keys.append(cell.value)
    return data

def convert_data(data1):
    data2 = json.dumps([dict(zip(map(lambda x:x['title'], data1["header"]), row)) for row in data1["data"]], separators=(',', ':'))
    data2_dict = json.loads(data2)
    print(data1)
    print(data2_dict)
    print(type(data2_dict))
    return data2_dict

def write_xlsx(file, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, key in enumerate(data[0].keys(), start=1):
        ws.cell(row=1, column=col, value=key)

    for row, item in enumerate(data, start=2):
        for col, value in enumerate(item.values(), start=1):
            ws.cell(row=row, column=col, value=value)
    
    wb.save(file)

def check_entity(name):
    if "CORP" in name or "LLC" in name or "LP" in name or "PC" in name or "CORPORATION" in name or "LLP" in name or "TRUST" in name or "REALTY" in name:
        return 1
    else:
        return 0

def build_url(data, entity_type, owner_num):
    if entity_type:
        address = data["Mailing Address"]
        citystatezip = data["Mailing City St  Zip"]
        search_url = "https://www.truepeoplesearch.com/resultaddress?streetaddress=" + address.replace(" ", "%20") + "&citystatezip=" + citystatezip.replace(" ", "%20")
    else:
        if owner_num == 1:
            name = data["Owner name 01"]
        if owner_num == 2:
            name = data["Owner name 02"]
        citystatezip = data["Mailing City St  Zip"]
        search_url = "https://www.truepeoplesearch.com/resultaddress?name=" + name.replace(" ", "%20") + "&citystatezip=" + citystatezip.replace(" ", "%20")

    return search_url


async def scrap(data, client, params):
    age_regex = re.compile(r'Age \d+ \([A-Za-z]{3} \d+\)')

    first_name = data["Owner name 01"]
    entity_type_1 = check_entity(first_name)

    search_url = build_url(data, entity_type_1, 1)

    print("\n")
    print(first_name)
    print(search_url)

    response = await client.get_async(search_url, params=params)

    with open("home_1.txt", "w", encoding="UTF-8") as output:
        output.write(response.text)

    soup = BeautifulSoup(response.text, 'html.parser')

    state = data["Mailing City St  Zip"].split()[0].lower()
    print(state)

    found_link = False
    link = ""

    for div in soup.find_all('div', {'class': 'card-summary'}):
        texts = div.text.split("\n")
        new_texts = [item.strip() for item in texts if item != '']
        
        ind = new_texts.index('Lives in')
        if state in new_texts[ind+1].lower():
            link = div.find('a', {'class': 'detail-link'})['href']
            break

    try:
        if link == "":
            link = soup.find('a', {'class': 'btn btn-success btn-lg detail-link shadow-form'})['href']
        detail_url = "https://www.truepeoplesearch.com" + link

        print(detail_url)

        detail_response = await client.get_async(detail_url, params=params)

        with open("detail_1.txt", "w", encoding="UTF-8") as output:
            output.write(detail_response.text)

        detail_soup = BeautifulSoup(detail_response.text, 'html.parser')
        
        owner = detail_soup.select_one('h1.oh1')
        if entity_type_1:
            data["Entity Owner's Name (Only if ENTITY)"] = owner.text
        
        try:
            age_span_1 = detail_soup.find('span', string=age_regex)    
            data["Age1"] = age_span_1.text.strip()
        except:
            data["Age1"] = "Age Unknown"
        
        phone_numbers_1 = []
        phone_divs = detail_soup.select('div.row div.col-12')
        for phone_div in phone_divs:
            try:
                phone_span = phone_div.select_one('a[data-link-to-more="phone"] span[itemprop="telephone"]')
                phone = phone_span.text.strip()

                phone_type_span = phone_div.select_one('span.smaller')
                phone_type = phone_type_span.text.strip()

                phone_info = f"{phone} - {phone_type}"
                if phone_info not in phone_numbers_1:
                    phone_numbers_1.append(phone_info)
            except:
                pass
        i = 1
        while len(phone_numbers_1):
            phone = phone_numbers_1.pop(0)
            if i < 7:
                data["Phone1-"+str(i)] = phone
                i += 1

        emails_1 = []
        email_divs = detail_soup.select('.col > div:last-child')
        for email_div in email_divs:
            text = email_div.text.strip()
            if re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text):
                if text != "support@truepeoplesearch.com" and text not in emails_1:
                    emails_1.append(text)

        j = 1
        while len(emails_1):
            email = emails_1.pop(0)
            if j < 3:
                data["Email1-"+str(j)] = email
                j += 1
    except:
        print("Cannot Find Detail Link")
        with open("error.txt", "a", encoding="UTF-8") as error:
            error.write(first_name + "\n")

    second_name = data["Owner name 02"]    
        
    if second_name:
        entity_type_2 = check_entity(second_name)
        if entity_type_1 == 1 and entity_type_2 == 1:
            pass
        else:
            search_url = build_url(data, entity_type_2, 2)

            print("\n\n")
            print(second_name)
            print(search_url)

            response = await client.get_async(search_url, params=params)
            with open("home2.txt", "w", encoding="UTF-8") as output:
                output.write(response.text)
            soup = BeautifulSoup(response.text, 'html.parser')

            link = ""
            for div in soup.find_all('div', {'class': 'card-summary'}):
                texts = div.text.split("\n")
                new_texts = [item.strip() for item in texts if item != '']
                
                ind = new_texts.index('Lives in')
                if state in new_texts[ind+1].lower():
                    link = div.find('a', {'class': 'detail-link'})['href']
                    break

            try:
                if link == "":
                    link = soup.find('a', {'class': 'btn btn-success btn-lg detail-link shadow-form'})['href']
                detail_url = "https://www.truepeoplesearch.com" + link

                print(detail_url)

                detail_response = await client.get_async(detail_url, params=params)

                with open("detail_2.txt", "w", encoding="UTF-8") as output:
                    output.write(detail_response.text)

                detail_soup = BeautifulSoup(detail_response.text, 'html.parser')
            
                try:
                    age_span_2 = detail_soup.find('span', string=age_regex)    
                    data["Age2"] = age_span_2.text.strip()
                except:
                    data["Age2"] = "Age Unknown"

                phone_numbers_2 = []
                phone_divs = detail_soup.select('div.row div.col-12')
                for phone_div in phone_divs:
                    try:
                        phone_span = phone_div.select_one('a[data-link-to-more="phone"] span[itemprop="telephone"]')
                        phone = phone_span.text.strip()

                        phone_type_span = phone_div.select_one('span.smaller')
                        phone_type = phone_type_span.text.strip()

                        phone_info = f"{phone} - {phone_type}"
                        if phone_info not in phone_numbers_2:
                            phone_numbers_2.append(phone_info)
                    except:
                        pass
                i = 1
                while len(phone_numbers_2):
                    phone = phone_numbers_2.pop(0)
                    if i < 7:
                        data["Phone2-"+str(i)] = phone
                        i += 1

                emails_2 = []
                email_divs = detail_soup.select('.col > div:last-child')
                for email_div in email_divs:
                    text = email_div.text.strip()
                    if re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text):
                        if text != "support@truepeoplesearch.com" and text not in emails_2:
                            emails_2.append(text)

                j = 1
                while len(emails_2):
                    email = emails_2.pop(0)
                    if j < 3:
                        data["Email2-"+str(j)] = email
                        j += 1
            except:
                print("Cannot Fine Detail Link")
                with open("error.txt", "a", encoding="UTF-8") as error:
                    error.write(second_name + "\n")

    return data

async def main(data):
    client = ZenRowsClient(zenrows_key, concurrency=10, retries=1)
    params = {"js_render":"true","antibot":"false","premium_proxy":"true"}

    # input_file = 'D:/Temp/BatchSkipTracing/Tracing_backend/Test.xlsx'
    input_data = convert_data(data)

    # Prompt for the range of rows to scrape
    # start_row = int(input("Enter the starting row (e.g., 1): "))
    # end_row = int(input("Enter the ending row (e.g., 20): "))

    # Slice the input_data based on the specified range
    input_temp = input_data
    
    output_data = []

    tasks = [scrap(item, client, params) for item in input_temp]
    output_data = await asyncio.gather(*tasks)

    print(output_data)

    # output_file = f"result{start_row:04d}-{end_row:04d}.xlsx"

    output_file = f"result_test.xlsx"
    
    write_xlsx(output_file, output_data)

    return {
        "Success": True,
        "data": output_data
    }


@router.post('/')
async def tracing_demo(req: Request):
    body = await req.body()
    data = json.loads(body)
    # try:
    #     result = asyncio.create_task(main(data))
    #     print(result)
    # except Exception as e:
    #    print("Error occurred:", str(e))
    print(data)
    return{
        "Success": True,
        "data": [{'File #': 1, 'Mailing Address': '139 BAY STREET', 'Mailing City St  Zip': 'STATEN ISLAND NY 10301', 'Owner name 01': '139 BAY POINTE PROPERTIES LLC', 'Owner name 02': '', "Entity Owner's Name (Only if ENTITY)": 'Jeanette Baldassano', 'Age1': 'Age 72 (Aug 1951)', 'Phone1-1': '(917) 474-6724 - Wireless', 'Phone1-2': '(718) 783-8789 - Landline', 'Phone1-3': '(718) 832-8057 - Landline', 'Phone1-4': '(607) 498-5311 - Landline', 'Phone1-5': '(718) 981-5817 - Landline', 'Phone1-6': '(718) 981-5316 - Landline', 'Email1-1': 'blennyb@aol.com'}, {'File #': '2', 'Mailing Address': '15 PAGE AVE', 'Mailing City St  Zip': 'STATEN ISLAND NY 10309-2611', 'Owner name 01': '139 BAY STREET REALTY LLC', 'Owner name 02': '', "Entity Owner's Name (Only if ENTITY)": 'Frank G Damato', 'Age1': 'Age 62 (Aug 1961)', 'Phone1-1': '(646) 808-6460 - Wireless', 'Phone1-2': '(718) 967-8393 - Landline', 'Phone1-3': '(276) 546-4274 - Landline', 'Phone1-4': '(732) 202-7154 - VOIP', 'Phone1-5': '(718) 763-9103 - Landline', 'Phone1-6': '(718) 967-0000 - Landline', 'Email1-1': 'fgd4712@aol.com'}]
    }
